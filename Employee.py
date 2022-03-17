import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1,column=1).value = "Employee ID"
sheet.cell(row=1,column=2).value = "Employee Name"
sheet.cell(row=1,column=3).value = "Basic Salary"
for i in range(2,12):		
    print("Enter the ",(i-1)," employee details")
    sheet.cell(row=i,column=1).value = int(input("Enter the emp id: "))
    sheet.cell(row=i,column=2).value = input("Enter the Employee Name: ")
    sheet.cell(row=i,column=3).value = int(input("Enter the Employee Basic Salary: "))
wb.save('/home/nischal/Python/Employee.xlsx')

wb1 = openpyxl.Workbook()
sheet1 = wb1.active
header = ['Employee ID', 'Employee Name', 'Basic Salary', 'DA', 'HRA', 'MA', 'TA', 'Gross Salary', 'EPF', 'Insurance', 'Net Salary']
for i in range(1,12):
	sheet1.cell(row=1,column=i).value = header[i-1]

print("Employee Id\tEmployee Name\tBasic Salary\t  DA\t\t\t  HRA\t\t\t  MA\t\t\t  TA\t\t\tGross Salary\t\t\t  Epf\t\t\t Insurance\t\tNet Salary\t")
row = []
for i in range(2,12):
    row.append(sheet.cell(row=i,column=1).value)
    row.append(sheet.cell(row=i,column=2).value)
    row.append(int(sheet.cell(row=i,column=3).value))
    row.append(0.8 * row[2])
    row.append(0.15 * row[2])
    row.append(0.12 * row[2])
    row.append(0.1 * row[2])
    row.append(sum(row[2:]))
    row.append(0.1133 * row[7])
    row.append(0.1 * row[7])
    row.append(row[7] - row[8] - row[9])
    print(row[0],"\t\t",row[1],"\t\t",row[2],"\t\t",row[3],"\t\t",row[4],"\t\t",row[5],"\t\t",row[6],"\t\t",row[7],f"\t\t\t{row[8]:.2f}","\t\t",row[9],f"\t\t{row[10]:.2f}")
    for j in range(1, 12):
    	sheet1.cell(row=i,column=j).value = row[j-1]
   
    row.clear()
wb1.save('/home/nischal/Python/Report.xlsx')
